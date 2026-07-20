from __future__ import annotations

import re
import json
import subprocess
from html import escape
from datetime import datetime
from io import BytesIO
from pathlib import Path
from typing import Iterable

import pandas as pd
import streamlit as st

try:
    import pdfplumber
except ImportError:
    pdfplumber = None


COMMON_ELEMENTS = {
    "c",
    "mn",
    "p",
    "s",
    "si",
    "al",
    "cr",
    "ni",
    "mo",
    "cu",
    "v",
    "nb",
    "ti",
    "b",
    "n",
    "ca",
    "co",
    "sn",
    "pb",
    "as",
    "w",
    "fe",
    "ce",
    "cev",
}

ELEMENT_ORDER = ["C", "Mn", "P", "S", "Si", "Ni", "Cr", "Mo", "Cu", "Al", "V", "Nb", "Ti", "N", "Pb", "B", "Co", "Sn"]
MECHANICAL_ORDER = [
    "Yield",
    "Tensile",
    "Elongation",
    "Charpy Temp F",
    "Charpy Temp C",
    "Charpy #1 FT-LB",
    "Charpy #2 FT-LB",
    "Charpy #3 FT-LB",
    "Charpy Avg FT-LB",
]

ELEMENT_ALIASES = {element: [element] for element in ELEMENT_ORDER}
MECHANICAL_ALIASES = {
    "Yield": ["yield", "yield ksi", "ys", "yield strength"],
    "Tensile": ["tensile", "uts", "tensile ksi", "tensile strength"],
    "Elongation": ["elongation", "elongation%", "elong", "el", "eu%"],
    "Charpy Temp F": ["charpy temperature f", "charpy temperature"],
    "Charpy Temp C": ["charpy temperature c", "charpy temperature"],
    "Charpy #1 FT-LB": ["charpy impact energy #1", "cvn ind1", "cvn #1"],
    "Charpy #2 FT-LB": ["charpy impact energy #2", "cvn ind2", "cvn #2"],
    "Charpy #3 FT-LB": ["charpy impact energy #3", "cvn ind3", "cvn #3"],
    "Charpy Avg FT-LB": ["charpy impact energy avg", "cvn avg"],
}
MECHANICAL_ALIASES = {label: [label, *aliases] for label, aliases in MECHANICAL_ALIASES.items()}
RULE_COLUMNS = ["Standard", "Grade", "Property", "Min", "Max", "Unit", "Note"]

PDF_FIELD_ALIASES = {
    **ELEMENT_ALIASES,
    **MECHANICAL_ALIASES,
    "Heat No": ["heat no", "heat number", "heat", "cast no", "cast number"],
    "FULL_TAG_NUM": ["full_tag_num", "full tag num", "coil no", "coil number", "coil", "tag no"],
    "BOL": ["bol", "b/l", "bill of lading", "bill lading"],
    "Material Spec": ["material spec", "material specification", "spec"],
}

IDENTIFIER_HINTS = {
    "bol",
    "heat",
    "dokum",
    "coil",
    "tag",
    "full",
    "id",
    "no",
    "number",
    "grade",
    "quality",
    "kalite",
    "size",
    "width",
    "thickness",
    "length",
    "weight",
    "vendor",
    "order",
    "line",
    "nbr",
    "so",
    "customer",
    "ship",
    "date",
}


def normalize_text(value: object) -> str:
    text = "" if value is None else str(value)
    text = text.strip().lower()
    replacements = str.maketrans("çğıöşüÇĞİÖŞÜ", "cgiosuCGIOSU")
    text = text.translate(replacements)
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def compact_text(value: object) -> str:
    return normalize_text(value).replace(" ", "")


def score_column(column: str, candidates: Iterable[str]) -> int:
    compact = compact_text(column)
    normalized = normalize_text(column)
    score = 0
    for candidate in candidates:
        candidate_compact = compact_text(candidate)
        if compact == candidate_compact:
            score += 100
        elif candidate_compact in compact:
            score += 40
        elif normalize_text(candidate) in normalized:
            score += 20
    return score


def best_column(columns: list[str], candidates: Iterable[str]) -> str | None:
    scored = [(score_column(column, candidates), column) for column in columns]
    scored = [item for item in scored if item[0] > 0]
    if not scored:
        return None
    return max(scored, key=lambda item: item[0])[1]


def ordered_detected_columns(columns: list[str], alias_map: dict[str, list[str]]) -> list[str]:
    detected = []
    for label in alias_map:
        column = best_column(columns, alias_map[label])
        if column and column not in detected:
            detected.append(column)
    return detected


def canonical_property_aliases(property_name: str) -> list[str]:
    name = str(property_name).strip()
    aliases = [name]
    aliases.extend(ELEMENT_ALIASES.get(name, []))
    aliases.extend(MECHANICAL_ALIASES.get(name, []))
    if name == "Nb":
        aliases.extend(["Cb", "Columbium", "Niobium"])
    if name == "Yield":
        aliases.extend(["Yield KSI", "Yield Strength", "Yield Strength KSI", "YS"])
    if name == "Tensile":
        aliases.extend(["Tensile KSI", "Tensile Strength", "Tensile Strength KSI", "UTS"])
    if name == "Elongation":
        aliases.extend(["Elongation%", "Total Elongation", "Elongation 2 in", "Elongation in 2 in", "EL"])
    return list(dict.fromkeys(aliases))


def resolve_property_column(columns: list[str], property_name: str) -> str | None:
    if property_name in columns:
        return property_name
    return best_column(columns, canonical_property_aliases(property_name))


def find_material_spec_column(df: pd.DataFrame) -> str | None:
    columns = list(df.columns)
    named_column = best_column(columns, ["material spec", "material specification", "spec", "standard", "grade", "quality", "kalite"])
    if named_column:
        return named_column

    best_match = None
    best_hits = 0
    for column in columns:
        sample = " ".join(df[column].dropna().astype(str).head(50).tolist())
        hits = len(re.findall(r"(?i)\bASTM\s*A\s*[0-9]+", sample))
        if hits > best_hits:
            best_hits = hits
            best_match = column
    return best_match if best_hits else None


def extract_pdf_content(uploaded_file) -> tuple[str, list[list[list[str]]]]:
    uploaded_file.seek(0)
    if pdfplumber is None:
        app_dir = Path(__file__).resolve().parent
        tmp_dir = app_dir / "tmp_pdf"
        tmp_dir.mkdir(exist_ok=True)
        pdf_path = tmp_dir / "uploaded.pdf"
        pdf_path.write_bytes(uploaded_file.read())

        bundled_python = (
            Path.home()
            / ".cache"
            / "codex-runtimes"
            / "codex-primary-runtime"
            / "dependencies"
            / "python"
            / "python.exe"
        )
        if not bundled_python.exists():
            raise RuntimeError("pdfplumber is not available for PDF reading. Run: pip install -r requirements.txt")

        script = (
            "import sys, json, pdfplumber\n"
            "parts=[]\n"
            "tables=[]\n"
            "with pdfplumber.open(sys.argv[1]) as pdf:\n"
            "    for page in pdf.pages:\n"
            "        parts.append(page.extract_text(x_tolerance=1, y_tolerance=3) or '')\n"
            "        for table in (page.extract_tables() or []):\n"
            "            clean=[]\n"
            "            for row in table:\n"
            "                clean.append([str(cell or '').strip() for cell in row])\n"
            "                parts.append(' '.join(str(cell or '') for cell in row))\n"
            "            tables.append(clean)\n"
            "print(json.dumps({'text':'\\n'.join(parts), 'tables': tables}, ensure_ascii=False))\n"
        )
        result = subprocess.run(
            [str(bundled_python), "-c", script, str(pdf_path)],
            capture_output=True,
            text=True,
            encoding="utf-8",
            errors="replace",
            check=False,
        )
        if result.returncode != 0:
            raise RuntimeError(result.stderr.strip() or "PDF text could not be extracted.")
        payload = json.loads(result.stdout)
        return payload["text"], payload["tables"]

    text_parts = []
    tables_out = []
    with pdfplumber.open(uploaded_file) as pdf:
        for page in pdf.pages:
            text_parts.append(page.extract_text(x_tolerance=1, y_tolerance=3) or "")
            tables = page.extract_tables() or []
            for table in tables:
                clean_table = []
                for row in table:
                    clean_row = [str(cell or "").strip() for cell in row]
                    clean_table.append(clean_row)
                    text_parts.append(" ".join(clean_row))
                tables_out.append(clean_table)
    return "\n".join(text_parts), tables_out


def extract_pdf_text(uploaded_file) -> str:
    text, _tables = extract_pdf_content(uploaded_file)
    return text


def extract_value_from_text(text: str, aliases: list[str], numeric_only: bool = True) -> str:
    for alias in aliases:
        escaped_alias = re.escape(alias).replace("\\ ", r"[\s_/-]*")
        if numeric_only:
            pattern = rf"(?i)\b{escaped_alias}\b\s*[:=\-]?\s*(-?\d+(?:[.,]\d+)?)"
        else:
            pattern = rf"(?i)\b{escaped_alias}\b\s*[:=\-]?\s*([A-Z0-9][A-Z0-9_\-/.]*)"
        match = re.search(pattern, text)
        if match:
            return match.group(1).replace(",", ".").strip()
    return ""


def extract_all_values_from_text(text: str, aliases: list[str]) -> list[str]:
    values = []
    for alias in aliases:
        escaped_alias = re.escape(alias).replace("\\ ", r"[\s_/-]*")
        pattern = rf"(?i)\b{escaped_alias}\b\s*[:=\-]?\s*([A-Z0-9][A-Z0-9_\-/.]*)"
        for match in re.finditer(pattern, text):
            value = match.group(1).strip()
            if value and value not in values:
                values.append(value)
    return values


def values_from_key_value_tables(tables: list[list[list[str]]], key_aliases: list[str]) -> list[str]:
    values = []
    wanted = {compact_text(alias) for alias in key_aliases}
    for table in tables:
        for row in table:
            for idx, cell in enumerate(row):
                if compact_text(cell) in wanted and idx + 1 < len(row):
                    value = str(row[idx + 1] or "").strip()
                    if value and value not in values:
                        values.append(value)
    return values


def value_from_key_value_tables(tables: list[list[list[str]]], key_aliases: list[str]) -> str:
    values = values_from_key_value_tables(tables, key_aliases)
    return values[0] if values else ""


def values_from_chemical_tables(tables: list[list[list[str]]]) -> dict[str, str]:
    values = {}
    wanted = {compact_text(element): element for element in ELEMENT_ORDER}
    for table in tables:
        for row_index, row in enumerate(table[:-1]):
            normalized_row = [compact_text(cell) for cell in row]
            element_hits = [wanted[cell] for cell in normalized_row if cell in wanted]
            if len(element_hits) < 4:
                continue

            value_row = table[row_index + 1]
            for col_index, cell in enumerate(row):
                element = wanted.get(compact_text(cell))
                if not element or col_index >= len(value_row):
                    continue
                raw_value = str(value_row[col_index] or "").strip().replace(",", ".")
                if re.fullmatch(r"-?\d+(?:\.\d+)?", raw_value):
                    values[element] = raw_value
    return values


def extract_sdi_mechanical_values(text: str) -> dict[str, str]:
    values = {}
    patterns = {
        "Yield": r"(?i)\bYield\s+Strength\s+(-?\d+(?:[.,]\d+)?)\s*KSI",
        "Tensile": r"(?i)\bTensile\s+Strength\s+(-?\d+(?:[.,]\d+)?)\s*KSI",
        "Elongation": r"(?i)\b(?:Total\s+)?Elongation\s+(-?\d+(?:[.,]\d+)?)\s*%",
        "Charpy Temp F": r"(?i)\bCharpy\s+Temperature\s+(-?\d+(?:[.,]\d+)?)\s*F",
        "Charpy Temp C": r"(?i)\bCharpy\s+Temperature\s+-?\d+(?:[.,]\d+)?\s*F\s+(-?\d+(?:[.,]\d+)?)\s*C",
    }

    for label, pattern in patterns.items():
        match = re.search(pattern, text)
        if match:
            values[label] = match.group(1).replace(",", ".").strip()

    pending_charpy_value = ""
    for line in text.splitlines():
        ft_lb_match = re.search(r"(-?\d+(?:[.,]\d+)?)\s*FT-LB", line, flags=re.IGNORECASE)
        if ft_lb_match:
            pending_charpy_value = ft_lb_match.group(1).replace(",", ".").strip()
            continue

        marker = line.strip().upper()
        if pending_charpy_value and marker in {"#1", "#2", "#3", "AVG"}:
            label = {
                "#1": "Charpy #1 FT-LB",
                "#2": "Charpy #2 FT-LB",
                "#3": "Charpy #3 FT-LB",
                "AVG": "Charpy Avg FT-LB",
            }[marker]
            values[label] = pending_charpy_value
            pending_charpy_value = ""

    return values


def pdf_to_dataframe(uploaded_file, fallback_bol: str = "") -> tuple[pd.DataFrame, str]:
    text, tables = extract_pdf_content(uploaded_file)
    record = {}

    bol = extract_value_from_text(text, PDF_FIELD_ALIASES["BOL"], numeric_only=False) or fallback_bol or "PDF"
    coil_values = values_from_key_value_tables(tables, ["Coil #", *PDF_FIELD_ALIASES["FULL_TAG_NUM"]])
    heat_values = values_from_key_value_tables(tables, ["Heat #", *PDF_FIELD_ALIASES["Heat No"]])

    if not coil_values:
        coil_values = extract_all_values_from_text(text, PDF_FIELD_ALIASES["FULL_TAG_NUM"])
    if not heat_values:
        heat_values = extract_all_values_from_text(text, PDF_FIELD_ALIASES["Heat No"])

    record["BOL"] = bol
    record["FULL_TAG_NUM"] = " | ".join(coil_values) if coil_values else ""
    record["Heat No"] = heat_values[0] if heat_values else "PDF_HEAT_1"
    record["Material Spec"] = value_from_key_value_tables(tables, PDF_FIELD_ALIASES["Material Spec"])

    chemical_values = values_from_chemical_tables(tables)
    for element in ELEMENT_ORDER:
        record[element] = chemical_values.get(element) or extract_value_from_text(text, PDF_FIELD_ALIASES[element], numeric_only=True)

    mechanical_values = extract_sdi_mechanical_values(text)
    for mechanical in MECHANICAL_ORDER:
        record[mechanical] = mechanical_values.get(mechanical) or extract_value_from_text(
            text,
            PDF_FIELD_ALIASES[mechanical],
            numeric_only=True,
        )

    if len(heat_values) > 1:
        rows = []
        for heat in heat_values:
            row = record.copy()
            row["Heat No"] = heat
            rows.append(row)
        df = pd.DataFrame(rows)
    else:
        df = pd.DataFrame([record])

    return clean_dataframe(df), text


def is_chemical_column(column: str, series: pd.Series) -> bool:
    normalized = normalize_text(column)
    compact = compact_text(column)
    tokens = set(normalized.split())

    if compact in COMMON_ELEMENTS or tokens.intersection(COMMON_ELEMENTS):
        return True

    if "%" in str(column) and not tokens.intersection(IDENTIFIER_HINTS):
        return True

    numeric_ratio = pd.to_numeric(series, errors="coerce").notna().mean()
    if numeric_ratio >= 0.75 and not any(hint in normalized for hint in IDENTIFIER_HINTS):
        return True

    return False


def clean_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    df.columns = [str(column).strip() for column in df.columns]
    df = df.dropna(how="all")
    df = df.dropna(axis=1, how="all")
    return df


def filter_by_bol(df: pd.DataFrame, bol_column: str, bol_number: str, use_contains: bool) -> pd.DataFrame:
    bol_values = df[bol_column].astype(str).map(normalize_text)
    needle = normalize_text(bol_number)
    if use_contains:
        mask = bol_values.str.contains(re.escape(needle), na=False)
    else:
        mask = bol_values == needle
    return df.loc[mask].copy()


def first_non_empty(values: pd.Series) -> object:
    non_empty = values.dropna()
    non_empty = non_empty[non_empty.astype(str).str.strip() != ""]
    if non_empty.empty:
        return ""
    return non_empty.iloc[0]


def distinct_join(values: pd.Series) -> str:
    cleaned = []
    for value in values.dropna():
        text = str(value).strip()
        if text and text not in cleaned:
            cleaned.append(text)
    return " | ".join(cleaned)


def build_heat_summary(
    filtered: pd.DataFrame,
    heat_column: str,
    coil_column: str,
    chemistry_columns: list[str],
) -> pd.DataFrame:
    grouped = filtered.groupby(heat_column, dropna=False, sort=True)
    summary = grouped[chemistry_columns].agg(first_non_empty).reset_index()
    summary.insert(1, "Coil Numbers", grouped[coil_column].agg(distinct_join).values)
    summary.insert(2, "Row Count", grouped.size().values)
    return summary


def find_conflicts(filtered: pd.DataFrame, heat_column: str, chemistry_columns: list[str]) -> pd.DataFrame:
    rows = []
    for heat, group in filtered.groupby(heat_column, dropna=False, sort=True):
        for column in chemistry_columns:
            values = [str(value).strip() for value in group[column].dropna() if str(value).strip()]
            distinct_values = []
            for value in values:
                if value not in distinct_values:
                    distinct_values.append(value)
            if len(distinct_values) > 1:
                rows.append({"Heat": heat, "Column": column, "Different Values": " | ".join(distinct_values)})
    return pd.DataFrame(rows)


def parse_float(value: object) -> float | None:
    if value is None or pd.isna(value):
        return None
    text = str(value).strip().replace(",", ".")
    match = re.search(r"-?\d+(?:\.\d+)?", text)
    if not match:
        return None
    return float(match.group(0))


def detect_standard_grade(text: object) -> tuple[str, str]:
    normalized = str(text or "")
    standard_match = re.search(r"(?i)\bASTM\s*A\s*([0-9]+)", normalized)
    grade_match = re.search(r"(?i)\b(?:GR|GRADE)\s*[-:/]?\s*([A-Z0-9]+)", normalized)
    class_match = re.search(r"(?i)\bCLASS\s*[-:/]?\s*([A-Z0-9]+)", normalized)
    type_match = re.search(r"(?i)\bTYPE\s*[-:/]?\s*([A-Z0-9]+)", normalized)
    standard = f"ASTM A{standard_match.group(1)}" if standard_match else ""
    grade = f"Grade {grade_match.group(1)}" if grade_match else ""
    if grade and class_match:
        grade = f"{grade} Class {class_match.group(1)}"
    elif grade and type_match:
        grade = f"{grade} Type {type_match.group(1)}"
    return standard, grade


def load_rules_from_file(uploaded_rules) -> pd.DataFrame:
    if uploaded_rules is None:
        rules_path = Path(__file__).resolve().parent / "standards_rules.csv"
        if rules_path.exists():
            rules = pd.read_csv(rules_path)
        else:
            rules = pd.DataFrame(columns=RULE_COLUMNS)
    elif uploaded_rules.name.lower().endswith(".csv"):
        rules = pd.read_csv(uploaded_rules)
    else:
        rules = pd.read_excel(uploaded_rules)

    for column in RULE_COLUMNS:
        if column not in rules.columns:
            rules[column] = ""
    return rules[RULE_COLUMNS].copy()


def sorted_rule_values(values: pd.Series) -> list[str]:
    cleaned = [str(value).strip() for value in values.dropna().tolist() if str(value).strip()]
    return sorted(set(cleaned), key=lambda value: [int(part) if part.isdigit() else part.lower() for part in re.split(r"(\d+)", value)])


def standard_options_from_rules(rules: pd.DataFrame, detected_standard: str = "") -> list[str]:
    options = sorted_rule_values(rules["Standard"]) if not rules.empty and "Standard" in rules.columns else []
    if detected_standard and detected_standard not in options:
        options.insert(0, detected_standard)
    return options


def grade_options_from_rules(rules: pd.DataFrame, standard: str, detected_grade: str = "") -> list[str]:
    if rules.empty or "Standard" not in rules.columns or "Grade" not in rules.columns:
        options = []
    else:
        standard_key = normalize_text(standard)
        matched = rules[rules["Standard"].astype(str).map(normalize_text) == standard_key]
        options = sorted_rule_values(matched["Grade"])
    if detected_grade and detected_grade not in options:
        options.insert(0, detected_grade)
    return options


def matching_rules(rules: pd.DataFrame, standard: str, grade: str) -> pd.DataFrame:
    if rules.empty:
        return rules
    standard_key = normalize_text(standard)
    grade_key = normalize_text(grade)
    base_grade_key = ""
    base_grade_match = re.search(r"(?i)\bgrade\s+[a-z0-9]+(?:\s+(?:class|type)\s+[a-z0-9]+)?", str(grade or ""))
    if base_grade_match:
        base_grade_key = normalize_text(base_grade_match.group(0))
    rule_standard = rules["Standard"].astype(str).map(normalize_text)
    rule_grade = rules["Grade"].astype(str).map(normalize_text)
    exact = rules[(rule_standard == standard_key) & ((rule_grade == grade_key) | (rule_grade == ""))].copy()
    if not exact.empty:
        return exact
    if base_grade_key and base_grade_key != grade_key:
        return rules[(rule_standard == standard_key) & ((rule_grade == base_grade_key) | (rule_grade == ""))].copy()
    return exact


def build_compliance_table(
    heat_summary: pd.DataFrame,
    heat_column: str,
    rules: pd.DataFrame,
    standard: str,
    grade: str,
) -> pd.DataFrame:
    active_rules = matching_rules(rules, standard, grade)
    rows = []

    if active_rules.empty:
        return pd.DataFrame(columns=["Heat No", "Coil / FULL_TAG_NUM", "Standard", "Grade", "Property", "Value", "Min", "Max", "Status", "Note"])

    for _, heat_row in heat_summary.iterrows():
        for _, rule in active_rules.iterrows():
            property_name = str(rule["Property"]).strip()
            value_column = resolve_property_column(list(heat_summary.columns), property_name)
            if value_column is None:
                status = "NO CHECK"
                value = ""
            else:
                value = heat_row[value_column]
                number = parse_float(value)
                min_value = parse_float(rule["Min"])
                max_value = parse_float(rule["Max"])
                if number is None:
                    status = "MISSING VALUE"
                elif min_value is not None and number < min_value:
                    status = "FAIL - BELOW MIN"
                elif max_value is not None and number > max_value:
                    status = "FAIL - ABOVE MAX"
                else:
                    status = "PASS"

            note = str(rule["Note"] or "")
            if status == "NO CHECK":
                note = (note + " | " if note else "") + f"No report column found for '{property_name}' or an equivalent alias."

            rows.append(
                {
                    "Heat No": heat_row[heat_column],
                    "Coil / FULL_TAG_NUM": heat_row["Coil Numbers"],
                    "Standard": standard,
                    "Grade": grade,
                    "Property": property_name,
                    "Value": value,
                    "Min": rule["Min"],
                    "Max": rule["Max"],
                    "Unit": rule["Unit"],
                    "Status": status,
                    "Note": note,
                }
            )

    return pd.DataFrame(rows)


def make_excel_report(
    bol_number: str,
    source_name: str,
    sheet_name: str,
    heat_summary: pd.DataFrame,
    element_summary: pd.DataFrame,
    mechanical_summary: pd.DataFrame,
    detail: pd.DataFrame,
    conflicts: pd.DataFrame,
    compliance: pd.DataFrame,
) -> bytes:
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter

        heat_column = heat_summary.columns[0]
        base_columns = [heat_column, "Coil Numbers", "Row Count"]
        element_columns = [column for column in element_summary.columns if column not in base_columns]
        mechanical_columns = [column for column in mechanical_summary.columns if column not in base_columns]
        report_columns = base_columns + element_columns + mechanical_columns
        report_df = heat_summary[report_columns].copy()
        report_df = report_df.rename(
            columns={
                heat_column: "Heat No",
                "Coil Numbers": "Coil / FULL_TAG_NUM",
                "Row Count": "Row Count",
            }
        )

        report_df.to_excel(writer, sheet_name="Report", startrow=3, index=False)
        if not compliance.empty:
            compliance.to_excel(writer, sheet_name="Compliance Check", index=False)
        detail.to_excel(writer, sheet_name="Detail", index=False)
        if not conflicts.empty:
            conflicts.to_excel(writer, sheet_name="Data Checks", index=False)

        workbook = writer.book
        worksheet = workbook["Report"]
        max_col = len(report_df.columns)
        max_row = len(report_df) + 4

        dark_fill = PatternFill("solid", fgColor="172033")
        blue_fill = PatternFill("solid", fgColor="26354D")
        teal_fill = PatternFill("solid", fgColor="2F5F73")
        light_fill = PatternFill("solid", fgColor="EEF3F7")
        border_color = "DBE2EA"
        thin_border = Border(
            left=Side(style="thin", color=border_color),
            right=Side(style="thin", color=border_color),
            top=Side(style="thin", color=border_color),
            bottom=Side(style="thin", color=border_color),
        )

        worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
        title_cell = worksheet.cell(row=1, column=1)
        title_cell.value = f"BOL {bol_number} - Heat / Coil Test Report"
        title_cell.fill = dark_fill
        title_cell.font = Font(color="FFFFFF", bold=True, size=16)
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        worksheet.row_dimensions[1].height = 28

        worksheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max_col)
        subtitle_cell = worksheet.cell(row=2, column=1)
        subtitle_cell.value = (
            f"Source: {source_name} | Sheet: {sheet_name} | "
            f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        )
        subtitle_cell.fill = light_fill
        subtitle_cell.font = Font(color="26354D", bold=True, size=10)
        subtitle_cell.alignment = Alignment(horizontal="center", vertical="center")

        for col in range(1, max_col + 1):
            worksheet.cell(row=3, column=col).border = thin_border
            worksheet.cell(row=3, column=col).alignment = Alignment(horizontal="center", vertical="center")

        worksheet.merge_cells(start_row=3, start_column=1, end_row=3, end_column=3)
        worksheet.cell(row=3, column=1).value = "Part Information"
        worksheet.cell(row=3, column=1).fill = blue_fill
        worksheet.cell(row=3, column=1).font = Font(color="FFFFFF", bold=True)

        element_start = 4
        element_end = element_start + len(element_columns) - 1
        mechanical_start = element_end + 1
        mechanical_end = mechanical_start + len(mechanical_columns) - 1

        if element_columns:
            worksheet.merge_cells(start_row=3, start_column=element_start, end_row=3, end_column=element_end)
            worksheet.cell(row=3, column=element_start).value = "Elements"
            worksheet.cell(row=3, column=element_start).fill = teal_fill
            worksheet.cell(row=3, column=element_start).font = Font(color="FFFFFF", bold=True)

        if mechanical_columns:
            worksheet.merge_cells(start_row=3, start_column=mechanical_start, end_row=3, end_column=mechanical_end)
            worksheet.cell(row=3, column=mechanical_start).value = "Mechanical Properties"
            worksheet.cell(row=3, column=mechanical_start).fill = blue_fill
            worksheet.cell(row=3, column=mechanical_start).font = Font(color="FFFFFF", bold=True)

        for cell in worksheet[4]:
            cell.fill = light_fill
            cell.font = Font(color="172033", bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

        for row in worksheet.iter_rows(min_row=5, max_row=max_row, max_col=max_col):
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center", vertical="center")
            row[0].font = Font(bold=True, color="172033")

        worksheet.freeze_panes = "D5"
        worksheet.auto_filter.ref = f"A4:{get_column_letter(max_col)}{max_row}"
        worksheet.sheet_view.showGridLines = False

        for col_idx, column_name in enumerate(report_df.columns, start=1):
            max_length = max(
                [len(str(column_name))]
                + [len(str(value)) for value in report_df.iloc[:, col_idx - 1].fillna("").head(200)]
            )
            worksheet.column_dimensions[get_column_letter(col_idx)].width = min(max(max_length + 2, 10), 28)
        worksheet.column_dimensions["B"].width = 34

        for worksheet in workbook.worksheets:
            if worksheet.title != "Report":
                worksheet.freeze_panes = "A2"
            for col_idx, column_cells in enumerate(worksheet.columns, start=1):
                max_length = max(len(str(cell.value or "")) for cell in column_cells)
                worksheet.column_dimensions[get_column_letter(col_idx)].width = min(max(max_length + 2, 10), 42)

    return output.getvalue()


def make_compliance_pdf_report(
    bol_number: str,
    source_name: str,
    standard: str,
    grade: str,
    heat_summary: pd.DataFrame,
    compliance: pd.DataFrame,
) -> bytes:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A3, landscape
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import inch
    from reportlab.platypus import PageBreak, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    output = BytesIO()
    doc = SimpleDocTemplate(
        output,
        pagesize=landscape(A3),
        leftMargin=0.35 * inch,
        rightMargin=0.35 * inch,
        topMargin=0.35 * inch,
        bottomMargin=0.35 * inch,
    )
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "ReportTitle",
        parent=styles["Title"],
        textColor=colors.HexColor("#172033"),
        fontSize=18,
        leading=22,
        spaceAfter=8,
    )
    section_style = ParagraphStyle(
        "SectionTitle",
        parent=styles["Heading2"],
        textColor=colors.HexColor("#172033"),
        fontSize=12,
        leading=15,
        spaceBefore=8,
        spaceAfter=6,
    )
    small_style = ParagraphStyle(
        "Small",
        parent=styles["BodyText"],
        fontSize=7,
        leading=9,
        textColor=colors.HexColor("#172033"),
    )
    small_bold_style = ParagraphStyle(
        "SmallBold",
        parent=small_style,
        fontName="Helvetica-Bold",
    )
    muted_style = ParagraphStyle(
        "Muted",
        parent=small_style,
        textColor=colors.HexColor("#475467"),
    )
    story = []

    failed = (
        compliance[
            compliance["Status"].astype(str).str.contains("FAIL|MISSING VALUE|NO CHECK", regex=True)
        ].copy()
        if not compliance.empty
        else pd.DataFrame()
    )

    def as_paragraph(value: object, style=small_style) -> Paragraph:
        return Paragraph(escape(str(value if value is not None else "")), style)

    source_files = [part.strip() for part in str(source_name or "").split(",") if part.strip()]
    if len(source_files) > 3:
        source_display = f"{len(source_files)} PDF files: " + ", ".join(source_files[:3]) + " ..."
    else:
        source_display = ", ".join(source_files) if source_files else "-"

    story.append(Paragraph(f"BOL {escape(str(bol_number or 'PDF_BATCH'))} - Standards Compliance Report", title_style))

    status_color = colors.HexColor("#DCFCE7") if failed.empty and not compliance.empty else colors.HexColor("#FEE2E2")
    status_text = "ALL CHECKS PASSED" if failed.empty and not compliance.empty else f"{len(failed)} FAILED / NEEDS REVIEW"
    badge = Table([[Paragraph(status_text, small_bold_style)]], colWidths=[2.5 * inch])
    badge.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, -1), status_color),
                ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#DBE2EA")),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("TOPPADDING", (0, 0), (-1, -1), 6),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
            ]
        )
    )
    story.append(badge)
    story.append(Spacer(1, 8))

    summary_rows = [
        ["Source", source_display],
        ["Standard", standard or "-"],
        ["Grade / Class", grade or "-"],
        ["Heat Count", str(heat_summary.iloc[:, 0].nunique() if not heat_summary.empty else 0)],
        ["Check Rows", str(len(compliance))],
        ["Failed / Needs Review", str(len(failed))],
        ["Generated At", datetime.now().strftime("%Y-%m-%d %H:%M")],
    ]
    summary_table = Table(
        [[as_paragraph(label, small_bold_style), as_paragraph(value, small_style)] for label, value in summary_rows],
        colWidths=[1.65 * inch, 13.55 * inch],
    )
    summary_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#EEF3F7")),
                ("TEXTCOLOR", (0, 0), (-1, -1), colors.HexColor("#172033")),
                ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#DBE2EA")),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("ROWBACKGROUNDS", (0, 0), (-1, -1), [colors.white, colors.HexColor("#F8FAFC")]),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ]
        )
    )
    story.append(summary_table)
    story.append(Spacer(1, 12))

    def status_fill(status: str):
        if "FAIL" in status:
            return colors.HexColor("#FEE2E2")
        if "MISSING VALUE" in status or "NO CHECK" in status:
            return colors.HexColor("#FEF3C7")
        return colors.HexColor("#DCFCE7")

    def add_table(title: str, frame: pd.DataFrame, max_rows: int | None = None) -> None:
        story.append(Paragraph(title, section_style))
        if frame.empty:
            story.append(Paragraph("Kayit yok.", muted_style))
            story.append(Spacer(1, 8))
            return

        display_columns = ["Heat No", "Coil / FULL_TAG_NUM", "Property", "Value", "Min", "Max", "Unit", "Status", "Note"]
        table_df = frame[[column for column in display_columns if column in frame.columns]].copy()
        if max_rows is not None:
            table_df = table_df.head(max_rows)

        rename_map = {
            "Heat No": "Heat",
            "Coil / FULL_TAG_NUM": "Coil",
            "Property": "Check",
            "Value": "Value",
            "Status": "Status",
            "Note": "Note",
        }
        headers = [rename_map.get(column, column) for column in table_df.columns]
        data = [[as_paragraph(header, small_bold_style) for header in headers]]
        for _, row in table_df.fillna("").astype(str).iterrows():
            data.append([as_paragraph(row[column], small_style) for column in table_df.columns])

        table = Table(
            data,
            repeatRows=1,
            colWidths=[
                0.85 * inch,
                2.35 * inch,
                1.0 * inch,
                0.65 * inch,
                0.55 * inch,
                0.55 * inch,
                0.45 * inch,
                1.55 * inch,
                6.25 * inch,
            ],
        )
        table_style = TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#172033")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#DBE2EA")),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8FAFC")]),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                ("LEFTPADDING", (0, 0), (-1, -1), 4),
                ("RIGHTPADDING", (0, 0), (-1, -1), 4),
            ]
        )
        status_col = list(table_df.columns).index("Status") if "Status" in table_df.columns else None
        if status_col is not None:
            for row_idx, status in enumerate(table_df["Status"].astype(str), start=1):
                table_style.add("BACKGROUND", (0, row_idx), (-1, row_idx), status_fill(status))
                table_style.add("FONTNAME", (status_col, row_idx), (status_col, row_idx), "Helvetica-Bold")
        table.setStyle(table_style)
        story.append(table)

        story.append(Spacer(1, 10))

    if compliance.empty:
        story.append(Paragraph("No standard compliance result was found. Enable the standard check and verify the Standard/Grade fields.", styles["Normal"]))
    else:
        add_table("Failed or Needs Review Values", failed)
        story.append(PageBreak())
        add_table("All Standard Check Results", compliance)

    doc.build(story)
    return output.getvalue()


def render_value_strip(row: pd.Series, columns: list[str]) -> str:
    chips = []
    for column in columns:
        value = row[column]
        display_value = "" if pd.isna(value) else value
        chips.append(
            '<div class="chem-chip">'
            f'<span>{escape(str(column))}</span>'
            f'<strong>{escape(str(display_value))}</strong>'
            '</div>'
        )
    return "".join(chips)


def render_chemistry_cards(
    summary: pd.DataFrame,
    heat_column: str,
    element_columns: list[str],
    mechanical_columns: list[str],
) -> None:
    for _, row in summary.iterrows():
        heat_value = row[heat_column]
        coil_values = row["Coil Numbers"]
        card_html = (
            '<section class="heat-card">'
            '<div class="heat-card-header">'
            '<div><div class="label">HEAT NO</div>'
            f'<div class="heat-no">{escape(str(heat_value))}</div></div>'
            '<div><div class="label">COIL / FULL_TAG_NUM</div>'
            f'<div class="coil-no">{escape(str(coil_values))}</div></div>'
            '</div>'
            '<div class="group-title">Elements</div>'
            f'<div class="chem-strip">{render_value_strip(row, element_columns)}</div>'
            '<div class="group-title">Mechanical Properties</div>'
            f'<div class="chem-strip mechanical-strip">{render_value_strip(row, mechanical_columns)}</div>'
            '</section>'
        )
        st.markdown(card_html, unsafe_allow_html=True)


st.set_page_config(page_title="BOL Heat Chemistry Lookup", page_icon="🔎", layout="wide")

st.markdown(
    """
    <style>
    .stApp {
        background: #f5f7fa;
    }
    .block-container {
        padding-top: 1.2rem;
        max-width: 1380px;
    }
    .hero {
        background: linear-gradient(135deg, #172033 0%, #26354d 58%, #2f5f73 100%);
        color: white;
        padding: 28px 30px;
        border-radius: 8px;
        margin-bottom: 18px;
        box-shadow: 0 14px 34px rgba(23, 32, 51, 0.16);
    }
    .hero h1 {
        margin: 0;
        font-size: 34px;
        letter-spacing: 0;
    }
    .hero p {
        margin: 8px 0 0 0;
        color: #d9e4ee;
        font-size: 15px;
    }
    .section-title {
        font-size: 20px;
        font-weight: 750;
        color: #172033;
        margin: 22px 0 10px;
    }
    .metric-card {
        background: white;
        border: 1px solid #dbe2ea;
        border-radius: 8px;
        padding: 16px 18px;
        box-shadow: 0 8px 22px rgba(23, 32, 51, 0.06);
    }
    .metric-card span {
        color: #667085;
        font-size: 13px;
    }
    .metric-card strong {
        display: block;
        color: #172033;
        font-size: 28px;
        margin-top: 4px;
    }
    .heat-card {
        background: white;
        border: 1px solid #dbe2ea;
        border-radius: 8px;
        margin-bottom: 12px;
        box-shadow: 0 8px 22px rgba(23, 32, 51, 0.06);
        overflow: hidden;
    }
    .heat-card-header {
        display: grid;
        grid-template-columns: minmax(220px, 0.75fr) minmax(280px, 1.25fr);
        gap: 18px;
        padding: 14px 18px;
        background: #eef3f7;
        border-bottom: 1px solid #dbe2ea;
    }
    .label {
        color: #667085;
        font-size: 12px;
        font-weight: 700;
    }
    .heat-no {
        color: #172033;
        font-size: 20px;
        font-weight: 800;
        margin-top: 3px;
    }
    .coil-no {
        color: #26354d;
        font-size: 14px;
        font-weight: 650;
        margin-top: 6px;
        overflow-wrap: anywhere;
    }
    .chem-strip {
        display: grid;
        grid-template-columns: repeat(auto-fill, minmax(92px, 1fr));
        gap: 8px;
        padding: 8px 14px 14px;
    }
    .mechanical-strip {
        grid-template-columns: repeat(auto-fill, minmax(150px, 1fr));
    }
    .group-title {
        color: #26354d;
        font-size: 13px;
        font-weight: 800;
        padding: 12px 14px 0;
        text-transform: uppercase;
        letter-spacing: 0;
    }
    .chem-chip {
        background: #f8fafc;
        border: 1px solid #dbe2ea;
        border-radius: 7px;
        padding: 8px 10px;
        min-width: 86px;
        min-height: 54px;
    }
    .chem-chip span {
        display: block;
        color: #667085;
        font-size: 11px;
        font-weight: 700;
    }
    .chem-chip strong {
        display: block;
        color: #172033;
        font-size: 15px;
        margin-top: 4px;
        overflow-wrap: anywhere;
    }
    div[data-testid="stDataFrame"] {
        border: 1px solid #dbe2ea;
        border-radius: 8px;
        overflow: hidden;
    }
    @media (max-width: 760px) {
        .hero h1 { font-size: 26px; }
        .heat-card-header { grid-template-columns: 1fr; }
    }
    </style>
    """,
    unsafe_allow_html=True,
)

st.markdown(
    """
    <div class="hero">
        <h1>BOL Heat Chemistry Lookup</h1>
        <p>Upload Excel or PDF files, enter a BOL number, and review chemistry, mechanical properties, and standard compliance by coil and heat.</p>
    </div>
    """,
    unsafe_allow_html=True,
)

uploaded_files = st.file_uploader(
    "Excel or PDF files",
    type=["xlsx", "xls", "xlsm", "pdf"],
    accept_multiple_files=True,
)

if not uploaded_files:
    st.info("Upload an Excel coil list or PDF test certificate to begin.")
    st.stop()

pdf_files = [file for file in uploaded_files if file.name.lower().endswith(".pdf")]
excel_files = [file for file in uploaded_files if not file.name.lower().endswith(".pdf")]
is_pdf = bool(pdf_files) and not excel_files
pdf_text = ""
source_name = ", ".join(file.name for file in uploaded_files)

if is_pdf:
    sheet_name = "PDF Batch" if len(pdf_files) > 1 else "PDF"
    pdf_frames = []
    pdf_text_parts = []
    pdf_errors = []
    for pdf_file in pdf_files:
        try:
            parsed_df, parsed_text = pdf_to_dataframe(pdf_file)
            parsed_df.insert(0, "Source PDF", pdf_file.name)
            pdf_frames.append(parsed_df)
            pdf_text_parts.append(f"--- {pdf_file.name} ---\n{parsed_text}")
        except Exception as exc:
            pdf_errors.append(f"{pdf_file.name}: {exc}")

    if pdf_errors:
        st.error("Some PDF files could not be read:\n" + "\n".join(pdf_errors))
        st.stop()

    if not pdf_frames:
        st.error("No readable PDF file was found.")
        st.stop()

    df = clean_dataframe(pd.concat(pdf_frames, ignore_index=True))
    pdf_text = "\n\n".join(pdf_text_parts)
else:
    if pdf_files and excel_files:
        st.error("Do not upload PDF and Excel files at the same time. Choose one file type for a single report.")
        st.stop()

    if len(excel_files) == 1:
        uploaded_file = excel_files[0]
        try:
            excel_file = pd.ExcelFile(uploaded_file)
        except Exception as exc:
            st.error(f"Excel file could not be read: {exc}")
            st.stop()

        sheet_name = st.selectbox("Sheet", excel_file.sheet_names)

        try:
            df = clean_dataframe(pd.read_excel(excel_file, sheet_name=sheet_name))
        except Exception as exc:
            st.error(f"The selected sheet could not be read: {exc}")
            st.stop()
    else:
        sheet_name = "Excel Batch"
        excel_frames = []
        excel_errors = []
        for excel_file_item in excel_files:
            try:
                parsed_excel = pd.ExcelFile(excel_file_item)
                first_sheet = parsed_excel.sheet_names[0]
                parsed_df = clean_dataframe(pd.read_excel(parsed_excel, sheet_name=first_sheet))
                parsed_df.insert(0, "Source Excel", excel_file_item.name)
                excel_frames.append(parsed_df)
            except Exception as exc:
                excel_errors.append(f"{excel_file_item.name}: {exc}")

        if excel_errors:
            st.error("Some Excel files could not be read:\n" + "\n".join(excel_errors))
            st.stop()
        if not excel_frames:
            st.error("No readable Excel file was found.")
            st.stop()

        df = clean_dataframe(pd.concat(excel_frames, ignore_index=True))

if df.empty:
    st.warning("No readable data was found in the selected sheet.")
    st.stop()

columns = list(df.columns)
detected_bol = best_column(columns, ["bol", "b/l", "bill of lading", "bill lading"])
detected_coil = best_column(columns, ["full_tag_num", "full tag num", "full tag", "coil", "coil no", "coil number"])
detected_heat = best_column(columns, ["heat", "heat no", "heat number", "dokum", "cast", "charge"])
detected_elements = ordered_detected_columns(columns, ELEMENT_ALIASES)
detected_mechanical = ordered_detected_columns(columns, MECHANICAL_ALIASES)
material_spec_column = find_material_spec_column(df)
detected_standard = ""
detected_grade = ""
if material_spec_column:
    material_values = df[material_spec_column].dropna().astype(str)
    for material_value in material_values.head(100):
        detected_standard, detected_grade = detect_standard_grade(material_value)
        if detected_standard:
            break

with st.sidebar:
    st.header("Standard Check")
    enable_standard_check = st.checkbox("Run standard limit check", value=False)
    uploaded_rules = st.file_uploader("Standard limit file", type=["csv", "xlsx", "xls"])
    standards_rules = load_rules_from_file(uploaded_rules)
    if standards_rules.empty:
        st.caption("No limit file was found. Example file: standards_rules.csv")
    else:
        st.caption(f"{len(standards_rules)} standard rule rows loaded.")

    standard_options = standard_options_from_rules(standards_rules, detected_standard)
    if standard_options:
        st.caption("Available standards: " + ", ".join(standard_options))
    standard_select_options = ["Select standard"] + standard_options + ["Custom"]
    default_standard_index = standard_select_options.index(detected_standard) if detected_standard in standard_select_options else 0
    selected_standard_option = st.selectbox(
        "Standard",
        standard_select_options,
        index=default_standard_index,
        help="Choose one of the ASTM standards loaded from the limit file.",
    )
    if selected_standard_option == "Custom":
        standard_name = st.text_input("Custom standard", value=detected_standard, placeholder="Example: ASTM A252")
    elif selected_standard_option == "Select standard":
        standard_name = ""
    else:
        standard_name = selected_standard_option

    grade_options = grade_options_from_rules(standards_rules, standard_name, detected_grade)
    grade_select_options = ["Select grade / class"] + grade_options + ["Custom"]
    default_grade_index = grade_select_options.index(detected_grade) if detected_grade in grade_select_options else 0
    selected_grade_option = st.selectbox(
        "Grade / Class",
        grade_select_options,
        index=default_grade_index,
        help="Choose the grade, class, or type covered by the selected ASTM standard.",
    )
    if selected_grade_option == "Custom":
        grade_name = st.text_input("Custom grade / class", value=detected_grade, placeholder="Example: Grade 3")
    elif selected_grade_option == "Select grade / class":
        grade_name = ""
    else:
        grade_name = selected_grade_option

    st.header("Column Settings")
    bol_column = st.selectbox(
        "BOL column",
        columns,
        index=columns.index(detected_bol) if detected_bol in columns else 0,
    )
    coil_column = st.selectbox(
        "Coil column",
        columns,
        index=columns.index(detected_coil) if detected_coil in columns else 0,
        help="For your files, this is usually the FULL_TAG_NUM column.",
    )
    heat_column = st.selectbox(
        "Heat column",
        columns,
        index=columns.index(detected_heat) if detected_heat in columns else 0,
    )
    material_options = ["None"] + columns
    material_spec_selection = st.selectbox(
        "Material Spec column",
        material_options,
        index=material_options.index(material_spec_column) if material_spec_column in material_options else 0,
        help="Select the column containing standard/grade information if it exists in the Excel file.",
    )
    element_columns = st.multiselect("Element columns", columns, default=detected_elements, key="element_columns_v2")
    mechanical_columns = st.multiselect(
        "Mechanical columns",
        columns,
        default=detected_mechanical,
        key="mechanical_columns_with_charpy",
    )
    use_contains = st.checkbox("Search BOL by contains", value=False)

    if is_pdf:
        with st.expander("Raw PDF text check"):
            st.text_area("Extracted PDF text", pdf_text[:5000], height=260)

bol_number = st.text_input(
    "BOL number",
    placeholder="Leave blank to list all uploaded rows/coils",
)

st.markdown('<div class="section-title">File Preview</div>', unsafe_allow_html=True)
preview_metrics = st.columns(3)
preview_metrics[0].markdown(f'<div class="metric-card"><span>Total rows</span><strong>{len(df)}</strong></div>', unsafe_allow_html=True)
preview_metrics[1].markdown(f'<div class="metric-card"><span>Total columns</span><strong>{len(df.columns)}</strong></div>', unsafe_allow_html=True)
preview_metrics[2].markdown(
    f'<div class="metric-card"><span>Report columns</span><strong>{len(element_columns) + len(mechanical_columns)}</strong></div>',
    unsafe_allow_html=True,
)
st.dataframe(df.head(15), use_container_width=True, hide_index=True)

report_columns = element_columns + [column for column in mechanical_columns if column not in element_columns]

if not element_columns:
    st.warning("Select at least one element column.")
    st.stop()

if not bol_number:
    filtered = df.copy()
else:
    filtered = filter_by_bol(df, bol_column, bol_number, use_contains)

if is_pdf and bol_number and filtered.empty:
    filtered = df.copy()
    filtered[bol_column] = bol_number

if filtered.empty:
    st.error("No rows were found for this BOL number. Check the selected column or search mode.")
    st.stop()

heat_summary = build_heat_summary(filtered, heat_column, coil_column, report_columns)
element_summary = heat_summary[[heat_column, "Coil Numbers", "Row Count"] + element_columns]
mechanical_summary = heat_summary[[heat_column, "Coil Numbers", "Row Count"] + mechanical_columns] if mechanical_columns else heat_summary[[heat_column, "Coil Numbers", "Row Count"]]
conflicts = find_conflicts(filtered, heat_column, report_columns)
if material_spec_selection != "None" and not filtered[material_spec_selection].dropna().empty:
    for material_value in filtered[material_spec_selection].dropna().astype(str).head(100):
        selected_standard, selected_grade = detect_standard_grade(material_value)
        if selected_standard:
            standard_name = standard_name or selected_standard
            grade_name = grade_name or selected_grade
            break
compliance = (
    build_compliance_table(heat_summary, heat_column, standards_rules, standard_name, grade_name)
    if enable_standard_check
    else pd.DataFrame()
)
failed_compliance = compliance[compliance["Status"].astype(str).str.contains("FAIL|MISSING VALUE|NO CHECK", regex=True)] if not compliance.empty else pd.DataFrame()
report_id = bol_number or "ALL"

st.markdown('<div class="section-title">Result Summary</div>', unsafe_allow_html=True)
result_metrics = st.columns(4)
result_metrics[0].markdown(f'<div class="metric-card"><span>BOL</span><strong>{report_id}</strong></div>', unsafe_allow_html=True)
result_metrics[1].markdown(f'<div class="metric-card"><span>Matched rows</span><strong>{len(filtered)}</strong></div>', unsafe_allow_html=True)
result_metrics[2].markdown(
    f'<div class="metric-card"><span>Heat count</span><strong>{filtered[heat_column].nunique(dropna=True)}</strong></div>',
    unsafe_allow_html=True,
)
result_metrics[3].markdown(
    f'<div class="metric-card"><span>Coil count</span><strong>{filtered[coil_column].nunique(dropna=True)}</strong></div>',
    unsafe_allow_html=True,
)

report_tab, table_tab, compliance_tab, detail_tab, control_tab = st.tabs(["Report View", "Heat Table", "Compliance Check", "Detail Rows", "Data Checks"])

with report_tab:
    st.markdown('<div class="section-title">Elements and Mechanical Properties by Heat</div>', unsafe_allow_html=True)
    render_chemistry_cards(heat_summary, heat_column, element_columns, mechanical_columns)

with table_tab:
    st.dataframe(heat_summary, use_container_width=True, hide_index=True)

with compliance_tab:
    if not enable_standard_check:
        st.info("Standard limit check is turned off.")
    elif compliance.empty:
        st.warning("No rules were found for this standard/grade, or the limit file is empty.")
    else:
        if failed_compliance.empty:
            st.success("All checked values are within limits.")
        else:
            st.error(f"{len(failed_compliance)} failed or needs-review rows found.")
            st.dataframe(failed_compliance, use_container_width=True, hide_index=True)
        st.dataframe(compliance, use_container_width=True, hide_index=True)

with detail_tab:
    display_columns = [bol_column, coil_column, heat_column]
    display_columns += [column for column in report_columns if column not in display_columns]
    remaining_columns = [column for column in filtered.columns if column not in display_columns]
    st.dataframe(filtered[display_columns + remaining_columns], use_container_width=True, hide_index=True)

with control_tab:
    if conflicts.empty:
        st.success("No conflicting values were found within the same heat.")
    else:
        st.warning("Different values were found within the same heat. Review before reporting.")
        st.dataframe(conflicts, use_container_width=True, hide_index=True)

report_bytes = make_excel_report(
    bol_number=report_id,
    source_name=source_name,
    sheet_name=sheet_name,
    heat_summary=heat_summary,
    element_summary=element_summary,
    mechanical_summary=mechanical_summary,
    detail=filtered,
    conflicts=conflicts,
    compliance=compliance,
)

pdf_report_bytes = make_compliance_pdf_report(
    bol_number=report_id,
    source_name=source_name,
    standard=standard_name,
    grade=grade_name,
    heat_summary=heat_summary,
    compliance=compliance,
)

download_cols = st.columns(2)
with download_cols[0]:
    st.download_button(
        "Download Excel report",
        data=report_bytes,
        file_name=f"BOL_{report_id}_heat_chemistry_report.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
with download_cols[1]:
    st.download_button(
        "Download PDF compliance report",
        data=pdf_report_bytes,
        file_name=f"BOL_{report_id}_compliance_report.pdf",
        mime="application/pdf",
    )
